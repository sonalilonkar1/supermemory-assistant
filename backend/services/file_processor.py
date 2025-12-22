"""File processing service for document and image uploads"""
import os
import uuid
import json
from typing import Dict, List, Optional, Tuple
from werkzeug.utils import secure_filename
import fitz  # PyMuPDF
from PIL import Image
import pytesseract
import docx
import openpyxl
import pandas as pd
from datetime import datetime, timezone

# Configure upload directory
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads')
ALLOWED_EXTENSIONS = {
    'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'docx', 'xlsx', 'xls', 'csv'
}

# Create upload directory if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename: str) -> bool:
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_uploaded_file(file, user_id: str) -> Optional[str]:
    """Save uploaded file and return file path"""
    if not file or not allowed_file(file.filename):
        return None
    
    # Generate unique filename
    filename = secure_filename(file.filename)
    unique_id = str(uuid.uuid4())
    extension = filename.rsplit('.', 1)[1].lower()
    new_filename = f"{user_id}_{unique_id}.{extension}"
    
    file_path = os.path.join(UPLOAD_FOLDER, new_filename)
    file.save(file_path)
    
    return file_path

def extract_text_from_file(file_path: str) -> Tuple[str, str]:
    """Extract text content from various file types"""
    filename = os.path.basename(file_path)
    extension = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    
    try:
        if extension == 'pdf':
            return extract_text_from_pdf(file_path), 'pdf'
        elif extension in ['png', 'jpg', 'jpeg', 'gif']:
            return extract_text_from_image(file_path), 'image'
        elif extension == 'docx':
            return extract_text_from_docx(file_path), 'docx'
        elif extension in ['xlsx', 'xls']:
            return extract_text_from_excel(file_path), 'excel'
        elif extension == 'csv':
            return extract_text_from_csv(file_path), 'csv'
        elif extension == 'txt':
            return extract_text_from_txt(file_path), 'txt'
        else:
            return f"Unsupported file type: {extension}", 'unknown'
    except Exception as e:
        return f"Error extracting text: {str(e)}", 'error'

def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF using PyMuPDF"""
    text = ""
    with fitz.open(file_path) as doc:
        for page in doc:
            text += page.get_text() + "\n"
    return text.strip()

def extract_text_from_image(file_path: str) -> str:
    """Extract text from image using OCR"""
    try:
        image = Image.open(file_path)
        text = pytesseract.image_to_string(image)
        return text.strip()
    except Exception as e:
        return f"OCR failed: {str(e)}"

def extract_text_from_docx(file_path: str) -> str:
    """Extract text from Word document"""
    doc = docx.Document(file_path)
    text = ""
    for paragraph in doc.paragraphs:
        text += paragraph.text + "\n"
    return text.strip()

def extract_text_from_excel(file_path: str) -> str:
    """Extract text from Excel file"""
    try:
        df = pd.read_excel(file_path)
        return df.to_string()
    except Exception as e:
        # Fallback to openpyxl
        workbook = openpyxl.load_workbook(file_path)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                text += "\t".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
        return text.strip()

def extract_text_from_csv(file_path: str) -> str:
    """Extract text from CSV file"""
    try:
        df = pd.read_csv(file_path)
        return df.to_string()
    except Exception as e:
        return "Error reading CSV file"

def extract_text_from_txt(file_path: str) -> str:
    """Extract text from plain text file"""
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read()

def create_attachment_metadata(file_path: str, extracted_text: str, file_type: str) -> Dict:
    """Create metadata for file attachment"""
    filename = os.path.basename(file_path)
    file_size = os.path.getsize(file_path)
    
    return {
        'filename': filename,
        'file_path': file_path,
        'file_type': file_type,
        'file_size': file_size,
        'extracted_text': extracted_text,  # Full text for memory creation
        'uploaded_at': datetime.now(timezone.utc).isoformat(),
        'text_length': len(extracted_text)
    }

def process_file_upload(file, user_id: str) -> Optional[Dict]:
    """Process uploaded file and return attachment metadata"""
    file_path = save_uploaded_file(file, user_id)
    if not file_path:
        return None
    
    extracted_text, file_type = extract_text_from_file(file_path)
    metadata = create_attachment_metadata(file_path, extracted_text, file_type)
    
    return metadata